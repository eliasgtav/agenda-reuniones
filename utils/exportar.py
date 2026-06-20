# © 2024 Elías Gaytan Alvino — Todos los derechos reservados.
import csv
import os
from datetime import datetime


def _ruta_descargas():
    # Carpeta Descargas en Windows (~\Downloads)
    downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
    os.makedirs(downloads, exist_ok=True)
    return downloads


def _ruta_salida(nombre):
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return os.path.join(_ruta_descargas(), f'{nombre}_{ts}')


def exportar_csv(reuniones, db):
    ruta = _ruta_salida('agenda_reuniones') + '.csv'
    with open(ruta, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            'ID', 'Asunto', 'Fecha', 'Hora', 'Lugar',
            'Estado', 'Participantes', 'Notas', 'Conclusión', 'Creado',
        ])
        for r in reuniones:
            parts = db.listar_participantes(r['id'])
            nombres = '; '.join(p['nombre'] for p in parts)
            writer.writerow([
                r['id'], r['asunto'], r['fecha'], r['hora'], r['lugar'],
                r['estado'], nombres, r['notas'], r['conclusion'], r['created_at'],
            ])
    return ruta


def exportar_excel(reuniones, db):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reuniones'

    COLORES_ESTADO = {
        'pendiente':   'FFFFF9C4',
        'realizada':   'FFC8E6C9',
        'cancelada':   'FFFFCDD2',
        'no_asistida': 'FFFFE0B2',
    }

    encabezados = ['ID', 'Asunto', 'Fecha', 'Hora', 'Lugar',
                   'Estado', 'Participantes', 'Notas', 'Conclusión', 'Creado']
    ws.append(encabezados)

    header_font = Font(bold=True, color='FFFFFFFF')
    header_fill = PatternFill('solid', fgColor='FF1565C0')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    anchos = [6, 40, 12, 8, 30, 12, 40, 50, 50, 20]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    for r in reuniones:
        parts = db.listar_participantes(r['id'])
        nombres = '; '.join(p['nombre'] for p in parts)
        fila = [
            r['id'], r['asunto'], r['fecha'], r['hora'], r['lugar'],
            r['estado'], nombres, r['notas'], r['conclusion'], r['created_at'],
        ]
        ws.append(fila)
        color = COLORES_ESTADO.get(r['estado'], 'FFFFFFFF')
        fill = PatternFill('solid', fgColor=color)
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ruta = _ruta_salida('agenda_reuniones') + '.xlsx'
    wb.save(ruta)
    return ruta
